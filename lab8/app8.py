import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Color

class CSVEntry:
    def __init__(self, line):
        values = self.prepare_line(line)
        self.show_id = values[0]
        self.type = values[1]
        self.title = values[2]
        self.director = values[3]
        self.cast = values[4].split(',')
        self.country = values[5]
        self.date_added = values[6]
        self.release_year = values[7]
        self.rating = values[8]
        self.duration = values[9]
        self.listed_in = values[10].split(',')

    def prepare_line(self, line):
        previous = 0
        inside_quotes = False
        values = []
        for i in range(len(line)):
            if(line[i] == ',' and not inside_quotes):
                values.append(line[previous:i])
                previous = i+1
            elif(line[i] == '"'):
                inside_quotes = not inside_quotes
        return values

#statistical
def get_average_duration(entries):
    sum = 0
    count = 0
    for entry in entries:
        if(entry.type == 'Movie' and 'min' in entry.duration):
            sum += int(entry.duration.split(' ')[0])
            count += 1
    return sum/count

#aggregation
def get_directors(entries):
    directors = {}
    for entry in entries:
        if entry.director in directors:
            directors[entry.director] += 1
        else:
            directors[entry.director] = 1
    return directors

#summarization
def get_total_number_of_tv_shows(entries):
        count = 0
        for entry in entries:
            if entry.type == 'TV Show':
                count += 1
        return count

MOVIES_BY_DIRECTOR = 'Movies by director'

def get_data(entries):
    data = {}
    data['Average time of movie in minutes'] = '{:.2f}'.format(get_average_duration(entries))
    data['Total number of TV Shows'] = get_total_number_of_tv_shows(entries)
    directors = get_directors(entries)
    sorted_directors = sorted(directors.items(), key=lambda x: x[1], reverse=True)
    data[MOVIES_BY_DIRECTOR] = sorted_directors[0:50]
    return data

def to_report(data, output):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    for i, key in enumerate(data):
        if(key==MOVIES_BY_DIRECTOR):
            ws.cell(row=i+1, column=1, value=key)
            for j, director in enumerate(data[key]):
                ws.cell(row=i+j+2, column=1, value=director[0])
                ws.cell(row=i+j+2, column=2, value=director[1])
                ws.cell(row=i+j+2, column=2).font = Font(color=Color("0000FF"))
        else:
            ws.cell(row=i+1, column=1, value=key)
            ws.cell(row=i+1, column=1).font = Font(bold=True)
            ws.cell(row=i+1, column=2, value=data[key])
            ws.cell(row=i+1, column=2).font = Font(name='Arial', size=12, color=Color("FF0000"))
            
    wb.save(output)

def to_console(data):
    for key in data:
        print(key + ': ' + str(data[key]))

def open_file(path):
    try:
        file = open(path, 'r', encoding='latin1')
    except FileNotFoundError:
        raise FileNotFoundError("File with given path does not exists")
    lines = file.readlines()
    file.close()
    entries = [CSVEntry(line) for line in lines[1:]]
    return entries


def run():
    parser = argparse.ArgumentParser(description='Process a dataset file.')
    parser.add_argument('filename', type=str, help='The path to the dataset file')
    parser.add_argument('-o', type=str, help='The file to store output', default='console')
    args = parser.parse_args()
    if(args.filename.endswith('.csv') == False):
        raise ValueError("File must be in CSV format")
    output = args.o
    entries = open_file(args.filename)
    data = get_data(entries)
    if output.endswith('.xlsx'):
        to_report(data, output)
    else:
        to_console(data)

if __name__ == "__main__":
    run()

